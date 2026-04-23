"""
智能表格识别与结构化提取模块
支持表格结构检测、单元格识别、数据导出
"""
import numpy as np
import cv2
from PIL import Image, ImageDraw
from typing import Dict, List, Tuple, Optional, Any
from dataclasses import dataclass
import json


@dataclass
class Cell:
    """表格单元格"""
    row: int
    col: int
    x: int
    y: int
    width: int
    height: int
    text: str = ""
    confidence: float = 0.0
    is_header: bool = False
    merged: bool = False
    merge_range: Optional[Tuple[int, int, int, int]] = None  # (start_row, end_row, start_col, end_col)


@dataclass
class Table:
    """表格数据结构"""
    cells: List[Cell]
    row_count: int
    col_count: int
    x: int
    y: int
    width: int
    height: int
    
    def to_dict(self) -> Dict:
        """转换为字典"""
        return {
            "row_count": self.row_count,
            "col_count": self.col_count,
            "position": {"x": self.x, "y": self.y, "width": self.width, "height": self.height},
            "cells": [
                {
                    "row": c.row,
                    "col": c.col,
                    "text": c.text,
                    "confidence": c.confidence,
                    "is_header": c.is_header,
                    "merged": c.merged,
                    "merge_range": c.merge_range
                }
                for c in self.cells
            ]
        }
    
    def to_2d_array(self) -> List[List[str]]:
        """转换为二维数组"""
        array = [["" for _ in range(self.col_count)] for _ in range(self.row_count)]
        for cell in self.cells:
            if 0 <= cell.row < self.row_count and 0 <= cell.col < self.col_count:
                array[cell.row][cell.col] = cell.text
        return array
    
    def to_markdown(self) -> str:
        """转换为Markdown表格"""
        array = self.to_2d_array()
        if not array:
            return ""
        
        lines = []
        # 表头
        lines.append("| " + " | ".join(array[0]) + " |")
        # 分隔符
        lines.append("| " + " | ".join(["---"] * len(array[0])) + " |")
        # 数据行
        for row in array[1:]:
            lines.append("| " + " | ".join(row) + " |")
        
        return "\n".join(lines)
    
    def to_html(self) -> str:
        """转换为HTML表格"""
        html = ['<table border="1" cellpadding="5" cellspacing="0">']
        
        array = self.to_2d_array()
        for i, row in enumerate(array):
            html.append("  <tr>")
            for j, cell_text in enumerate(row):
                tag = "th" if i == 0 else "td"
                html.append(f"    <{tag}>{cell_text}</{tag}>")
            html.append("  </tr>")
        
        html.append("</table>")
        return "\n".join(html)


class TableExtractor:
    """
    表格提取器
    基于图像处理和OCR的表格结构识别
    """
    
    def __init__(self):
        self.min_table_area = 10000  # 最小表格面积
        self.line_threshold = 150    # 直线检测阈值
    
    def extract_tables(self, image: Image.Image, ocr_result: Dict) -> List[Table]:
        """
        从图像中提取所有表格
        
        Args:
            image: PIL图像
            ocr_result: OCR识别结果，包含文本框信息
        
        Returns:
            表格列表
        """
        img_array = np.array(image)
        gray = cv2.cvtColor(img_array, cv2.COLOR_RGB2GRAY)
        
        # 1. 检测表格区域
        table_regions = self._detect_table_regions(gray)
        
        tables = []
        for region in table_regions:
            # 2. 提取表格结构
            table = self._extract_table_structure(gray, region, ocr_result)
            if table:
                tables.append(table)
        
        return tables
    
    def _detect_table_regions(self, gray: np.ndarray) -> List[Tuple[int, int, int, int]]:
        """检测图像中的表格区域"""
        # 边缘检测
        edges = cv2.Canny(gray, 50, 150)
        
        # 膨胀操作，连接断开的线条
        kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (5, 5))
        edges = cv2.dilate(edges, kernel, iterations=1)
        
        # 查找轮廓
        contours, _ = cv2.findContours(edges, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        
        regions = []
        for contour in contours:
            area = cv2.contourArea(contour)
            if area < self.min_table_area:
                continue
            
            x, y, w, h = cv2.boundingRect(contour)
            aspect_ratio = w / h if h > 0 else 0
            
            # 表格通常有合理的宽高比
            if 0.2 < aspect_ratio < 5:
                regions.append((x, y, w, h))
        
        return regions
    
    def _extract_table_structure(self, gray: np.ndarray, 
                                  region: Tuple[int, int, int, int],
                                  ocr_result: Dict) -> Optional[Table]:
        """提取表格结构"""
        x, y, w, h = region
        
        # 提取区域图像
        roi = gray[y:y+h, x:x+w]
        
        # 1. 检测水平和垂直线
        horizontal_lines, vertical_lines = self._detect_lines(roi)
        
        if len(horizontal_lines) < 2 or len(vertical_lines) < 2:
            return None
        
        # 2. 计算单元格位置
        rows = self._calculate_rows(horizontal_lines, h)
        cols = self._calculate_cols(vertical_lines, w)
        
        row_count = len(rows) - 1
        col_count = len(cols) - 1
        
        if row_count < 1 or col_count < 1:
            return None
        
        # 3. 获取OCR文本框
        text_boxes = ocr_result.get('boxes', [])
        texts = ocr_result.get('texts', [])
        confidences = ocr_result.get('confidences', [])
        
        # 4. 创建单元格并分配文本
        cells = []
        for i in range(row_count):
            for j in range(col_count):
                cell_x = x + cols[j]
                cell_y = y + rows[i]
                cell_w = cols[j+1] - cols[j]
                cell_h = rows[i+1] - rows[i]
                
                # 查找落在该单元格内的文本
                cell_text, cell_conf = self._get_cell_text(
                    cell_x, cell_y, cell_w, cell_h,
                    text_boxes, texts, confidences
                )
                
                cell = Cell(
                    row=i,
                    col=j,
                    x=cell_x,
                    y=cell_y,
                    width=cell_w,
                    height=cell_h,
                    text=cell_text,
                    confidence=cell_conf,
                    is_header=(i == 0)  # 第一行默认为表头
                )
                cells.append(cell)
        
        # 5. 检测合并单元格
        cells = self._detect_merged_cells(cells, row_count, col_count)
        
        return Table(
            cells=cells,
            row_count=row_count,
            col_count=col_count,
            x=x,
            y=y,
            width=w,
            height=h
        )
    
    def _detect_lines(self, roi: np.ndarray) -> Tuple[List[int], List[int]]:
        """检测表格线"""
        # 二值化
        _, binary = cv2.threshold(roi, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
        
        # 检测水平线
        horizontal_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (roi.shape[1] // 4, 1))
        horizontal_lines_img = cv2.morphologyEx(binary, cv2.MORPH_OPEN, horizontal_kernel)
        
        horizontal_lines = []
        for y in range(horizontal_lines_img.shape[0]):
            if np.sum(horizontal_lines_img[y, :]) > roi.shape[1] * 0.3:
                horizontal_lines.append(y)
        
        # 检测垂直线
        vertical_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, roi.shape[0] // 4))
        vertical_lines_img = cv2.morphologyEx(binary, cv2.MORPH_OPEN, vertical_kernel)
        
        vertical_lines = []
        for x in range(vertical_lines_img.shape[1]):
            if np.sum(vertical_lines_img[:, x]) > roi.shape[0] * 0.3:
                vertical_lines.append(x)
        
        # 合并接近的线
        horizontal_lines = self._merge_close_lines(horizontal_lines, 10)
        vertical_lines = self._merge_close_lines(vertical_lines, 10)
        
        return horizontal_lines, vertical_lines
    
    def _merge_close_lines(self, lines: List[int], threshold: int) -> List[int]:
        """合并接近的线条"""
        if not lines:
            return lines
        
        lines = sorted(lines)
        merged = [lines[0]]
        
        for line in lines[1:]:
            if line - merged[-1] > threshold:
                merged.append(line)
            else:
                merged[-1] = (merged[-1] + line) // 2
        
        return merged
    
    def _calculate_rows(self, horizontal_lines: List[int], height: int) -> List[int]:
        """计算行边界"""
        rows = [0] + horizontal_lines + [height]
        return sorted(list(set(rows)))
    
    def _calculate_cols(self, vertical_lines: List[int], width: int) -> List[int]:
        """计算列边界"""
        cols = [0] + vertical_lines + [width]
        return sorted(list(set(cols)))
    
    def _get_cell_text(self, cell_x: int, cell_y: int, cell_w: int, cell_h: int,
                       text_boxes: List, texts: List, confidences: List) -> Tuple[str, float]:
        """获取单元格内的文本"""
        cell_texts = []
        cell_confs = []
        
        for box, text, conf in zip(text_boxes, texts, confidences):
            # 计算文本框中心点
            if isinstance(box, list) and len(box) >= 4:
                # 四点坐标格式 [[x1,y1], [x2,y2], [x3,y3], [x4,y4]]
                center_x = sum(p[0] for p in box) / 4
                center_y = sum(p[1] for p in box) / 4
            elif isinstance(box, dict):
                center_x = box.get('x', 0) + box.get('width', 0) / 2
                center_y = box.get('y', 0) + box.get('height', 0) / 2
            else:
                continue
            
            # 检查中心点是否在单元格内
            if (cell_x <= center_x <= cell_x + cell_w and
                cell_y <= center_y <= cell_y + cell_h):
                cell_texts.append(text)
                cell_confs.append(conf if isinstance(conf, (int, float)) else 0.0)
        
        if cell_texts:
            return " ".join(cell_texts), np.mean(cell_confs)
        return "", 0.0
    
    def _detect_merged_cells(self, cells: List[Cell], row_count: int, col_count: int) -> List[Cell]:
        """检测合并单元格"""
        # 创建单元格查找表
        cell_map = {}
        for cell in cells:
            cell_map[(cell.row, cell.col)] = cell
        
        # 检测水平合并（跨列）
        for i in range(row_count):
            j = 0
            while j < col_count:
                cell = cell_map.get((i, j))
                if cell and not cell.text:
                    # 检查是否跨列合并
                    merge_end = j + 1
                    while merge_end < col_count:
                        next_cell = cell_map.get((i, merge_end))
                        if next_cell and next_cell.text:
                            break
                        merge_end += 1
                    
                    if merge_end > j + 1:
                        # 标记合并单元格
                        cell.merged = True
                        cell.merge_range = (i, i, j, merge_end - 1)
                        j = merge_end
                        continue
                j += 1
        
        # 检测垂直合并（跨行）
        for j in range(col_count):
            i = 0
            while i < row_count:
                cell = cell_map.get((i, j))
                if cell and not cell.text:
                    # 检查是否跨行合并
                    merge_end = i + 1
                    while merge_end < row_count:
                        next_cell = cell_map.get((merge_end, j))
                        if next_cell and next_cell.text:
                            break
                        merge_end += 1
                    
                    if merge_end > i + 1:
                        cell.merged = True
                        if cell.merge_range:
                            cell.merge_range = (i, merge_end - 1, cell.merge_range[2], cell.merge_range[3])
                        else:
                            cell.merge_range = (i, merge_end - 1, j, j)
                        i = merge_end
                        continue
                i += 1
        
        return cells
    
    def visualize_tables(self, image: Image.Image, tables: List[Table]) -> Image.Image:
        """可视化表格检测结果"""
        vis_image = image.copy()
        draw = ImageDraw.Draw(vis_image)
        
        colors = ['red', 'blue', 'green', 'orange', 'purple']
        
        for idx, table in enumerate(tables):
            color = colors[idx % len(colors)]
            
            # 绘制表格边框
            draw.rectangle(
                [table.x, table.y, table.x + table.width, table.y + table.height],
                outline=color,
                width=3
            )
            
            # 绘制单元格
            for cell in table.cells:
                draw.rectangle(
                    [cell.x, cell.y, cell.x + cell.width, cell.y + cell.height],
                    outline=color,
                    width=1
                )
                
                # 标记表头
                if cell.is_header:
                    draw.rectangle(
                        [cell.x, cell.y, cell.x + cell.width, cell.y + cell.height],
                        fill=(255, 200, 200, 128)
                    )
        
        return vis_image


class TableExporter:
    """表格导出器"""
    
    @staticmethod
    def to_excel(tables: List[Table], filepath: str):
        """导出为Excel文件"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, Border, Side
            
            wb = openpyxl.Workbook()
            
            for idx, table in enumerate(tables):
                if idx == 0:
                    ws = wb.active
                    ws.title = f"Table_{idx+1}"
                else:
                    ws = wb.create_sheet(f"Table_{idx+1}")
                
                # 写入数据
                array = table.to_2d_array()
                for i, row in enumerate(array):
                    for j, value in enumerate(row):
                        cell = ws.cell(row=i+1, column=j+1, value=value)
                        
                        # 表头样式
                        if i == 0:
                            cell.font = Font(bold=True)
                            cell.fill = openpyxl.styles.PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                        
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # 调整列宽
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column].width = adjusted_width
            
            wb.save(filepath)
            return True
        except ImportError:
            print("[TableExporter] openpyxl not installed, cannot export to Excel")
            return False
    
    @staticmethod
    def to_json(tables: List[Table], filepath: str):
        """导出为JSON文件"""
        data = {
            "table_count": len(tables),
            "tables": [table.to_dict() for table in tables]
        }
        
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        
        return True
    
    @staticmethod
    def to_csv(tables: List[Table], filepath: str):
        """导出为CSV文件"""
        import csv
        
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            if not tables:
                return True
            
            writer = csv.writer(f)
            array = tables[0].to_2d_array()
            writer.writerows(array)
        
        return True


# 全局提取器实例
_extractor: Optional[TableExtractor] = None


def get_table_extractor() -> TableExtractor:
    """获取全局表格提取器实例"""
    global _extractor
    if _extractor is None:
        _extractor = TableExtractor()
    return _extractor


def extract_tables(image: Image.Image, ocr_result: Dict) -> List[Table]:
    """
    提取表格的便捷函数
    
    Args:
        image: PIL图像
        ocr_result: OCR识别结果
    
    Returns:
        表格列表
    """
    extractor = get_table_extractor()
    return extractor.extract_tables(image, ocr_result)
